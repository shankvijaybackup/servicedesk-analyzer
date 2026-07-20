"""Excel workbook of the aggregated analysis (no raw ticket rows, so no PII)."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_HEAD_FILL = PatternFill("solid", fgColor="1F2937")
_HEAD_FONT = Font(color="FFFFFF", bold=True)


def write(a: dict, path) -> str:
    wb = Workbook()
    wb.remove(wb.active)

    dq = a["data_quality"]
    _sheet(wb, "Summary", [["Field", "Value"]] + [
        ["Source", a["meta"]["source_name"]],
        ["Generated", a["meta"]["generated_at"]],
        ["Tool", f"{a['meta']['tool']} v{a['meta']['version']}"],
        ["Total records", dq["total_records"]],
        ["Data quality", f"{dq['quality_grade']} ({dq['quality_score']}/100)"],
        ["Est. deflectable %", "-".join(map(str, a["opportunities"]["roi_summary"]["est_total_deflectable_pct_range"]))],
        ["Note", a["meta"]["stateless_note"]],
    ])

    _sheet(wb, "DataQuality", [["Metric", "Value"]] + [
        ["Fields detected", ", ".join(dq["fields_detected"])],
        ["Fields missing", ", ".join(dq["fields_missing"])],
        ["Duplicates", dq["duplicate_records"]],
        ["Duplicate basis", dq["duplicate_basis"]],
        ["MTTR available", dq["mttr_available"]],
        ["MTTR source", dq["mttr_source"]],
        ["Quality reasons", "; ".join(dq["quality_reasons"])],
    ])

    vol = a["volume"]
    vol_rows = [["Dimension", "Value", "Count", "Pct"]]
    for key in ("by_type", "by_priority", "by_status"):
        for r in vol.get(key, []):
            vol_rows.append([key.replace("by_", ""), r["value"], r["count"], r.get("pct", "")])
    for r in vol.get("by_month", []):
        vol_rows.append(["month", r["month"], r["count"], ""])
    _sheet(wb, "Volume", vol_rows)

    m = a["mttr"]
    mttr_rows = [["Scope", "Value", "n", "Median (h)", "Mean (h)", "P90 (h)"]]
    if m.get("available"):
        o = m["overall"]
        mttr_rows.append(["overall", "", o["n"], o["median_hours"], o["mean_hours"], o["p90_hours"]])
        for r in m.get("by_theme", []):
            mttr_rows.append(["theme", r["value"], r["n"], r["median_hours"], r["mean_hours"], r["p90_hours"]])
    _sheet(wb, "MTTR", mttr_rows)

    _sheet(wb, "Themes", [["Theme", "Count", "Pct", "MTTR median (h)", "MTTR p90 (h)",
                           "Description coverage %", "Confidence"]] + [
        [t["theme"], t["count"], t["pct"], t["mttr_median_hours"], t["mttr_p90_hours"],
         t["description_coverage_pct"], t["confidence"]] for t in a["themes"]])

    al = a["application_landscape"]
    app_rows = [["Application", "Count", "Pct"]]
    if al and al.get("available"):
        for r in al["top"]:
            app_rows.append([r["application"], r["count"], r["pct"]])
    _sheet(wb, "Applications", app_rows)

    opp = a["opportunities"]
    _sheet(wb, "OpportunityBacklog", [["Theme", "Primary type", "Addressable", "Deflect low",
                                       "Deflect high", "MTTR impact", "Complexity", "Risk",
                                       "Confidence", "Capabilities"]] + [
        [b["theme"], b["primary_type"], b["tickets_addressable"],
         b["est_deflectable_tickets_range"][0], b["est_deflectable_tickets_range"][1],
         b["mttr_reduction_potential"], b["implementation_complexity"], b["risk_level"],
         b["confidence"], ", ".join(b["atomicwork_capabilities"])] for b in opp["backlog"]])

    _sheet(wb, "AgenticBacklog", [["Theme", "Tickets", "Trigger", "System of action",
                                   "Permissions", "Feasibility", "Risk", "Human approval",
                                   "Expected impact"]] + [
        [g["theme"], g["tickets_addressable"], g["trigger"], g["system_of_action"],
         g["required_permissions"], g["automation_feasibility"], g["risk_level"],
         g["human_approval_required"], g["expected_impact"]] for g in opp["agentic_backlog"]])

    _sheet(wb, "SolutionMap", [["Theme", "Tickets", "Primary type", "Capabilities"]] + [
        [s["theme"], s["tickets"], s["primary_type"], ", ".join(s["capabilities"])]
        for s in opp["solution_map"]])

    impl = a.get("implementation")
    if impl:
        tp = impl["test_plan"]
        _sheet(wb, "UATTestPlan", [["ID", "Area", "Test case", "Role", "Priority",
                                    "Steps", "Expected result", "Derived from",
                                    "Status", "Tester", "Date", "Defect ref", "Notes"]] + [
            [c["id"], c["area"], c["title"], c["role"], c["priority"],
             " | ".join(c["steps"]), c["expected"], c["source"], "", "", "", "", ""]
            for c in tp["cases"]])

        _sheet(wb, "RoleMatrix", [["Role", "Can", "Cannot", "Derived from", "Verified"]] + [
            [r["role"], "; ".join(r["can"]), "; ".join(r["cannot"]), r["derived_from"], ""]
            for r in impl["role_matrix"]])

        _sheet(wb, "RACI", [["Activity", "Responsible", "Accountable", "Consulted",
                             "Informed"]] + [
            [r["activity"], r["responsible"], r["accountable"], r["consulted"],
             r["informed"]] for r in impl["raci"]])

        _sheet(wb, "Readiness", [["Gate", "Category", "Item", "Derived from",
                                  "Done", "Sign-off"]] + [
            [r["gate"], r["category"], r["item"], r["source"], "", ""]
            for r in impl["readiness_checklist"]])

        _sheet(wb, "PhasePlan", [["Days", "Phase", "Activities", "Exit criteria"]] + [
            [p["days"], p["phase"], " | ".join(p["activities"]), p["exit_criteria"]]
            for p in impl["phase_plan_15_day"]])

    iteration = a.get("iteration")
    if iteration:
        rows = [["Metric", "Baseline", "Follow-up", "Absolute change", "Improvement %",
                 "Baseline n", "Follow-up n"]]
        for metric, change in iteration["changes"].items():
            before = iteration["baseline"]["metrics"][metric]
            after = iteration["follow_up"]["metrics"][metric]
            rows.append([metric, before["value"], after["value"], change.get("absolute"),
                         change.get("improvement_pct"), before["denominator"], after["denominator"]])
        _sheet(wb, "IterationScorecard", rows)
        _sheet(wb, "IterationDecision", [["Field", "Value"],
            ["Pilot", iteration["pilot"]["name"]],
            ["Decision", iteration["decision"]["code"]],
            ["Reasons", "; ".join(iteration["decision"]["reasons"])],
            ["Comparability", iteration["comparability"]["status"]],
            ["Method", iteration["method"]],
            ["Causality note", iteration["causality_note"]],
        ])

    path = Path(path)
    wb.save(path)
    return str(path)


def _sheet(wb, title, rows):
    ws = wb.create_sheet(title[:31])
    for r in rows:
        ws.append([_cell(v) for v in r])
    for c in range(1, len(rows[0]) + 1):
        ws.cell(row=1, column=c).fill = _HEAD_FILL
        ws.cell(row=1, column=c).font = _HEAD_FONT
    _autofit(ws, rows)
    ws.freeze_panes = "A2"
    return ws


def _cell(v):
    if isinstance(v, (list, tuple)):
        v = ", ".join(map(str, v))
    if v is None:
        return ""
    # Excel interprets strings beginning with these characters as formulas.
    # Report labels come from uploaded data, so force them to remain literal
    # text. Numeric values (including legitimate negative numbers) are left
    # untouched because this branch applies only to strings.
    if isinstance(v, str) and v.startswith(("=", "+", "-", "@")):
        return "'" + v
    return v


def _autofit(ws, rows):
    ncols = len(rows[0])
    for c in range(1, ncols + 1):
        width = 12
        for r in rows:
            if c - 1 < len(r):
                width = max(width, min(60, len(str(_cell(r[c - 1]))) + 2))
        ws.column_dimensions[get_column_letter(c)].width = width
        for row in range(1, len(rows) + 1):
            ws.cell(row=row, column=c).alignment = Alignment(vertical="top", wrap_text=True)
